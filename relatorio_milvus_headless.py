# -*- coding: utf-8 -*-
"""
relatorio_milvus.py
Script para gerar relatórios diários de atendimento (Milvus), enviar e-mail com anexo,
alertar técnicos via WhatsApp e atualizar a planilha mensal de horas.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Tuple
import io
import logging
import smtplib
import socket
import time
import os # Importar os para deletar o arquivo
import json

import pandas as pd
import pywhatkit as kit
import requests
from email.message import EmailMessage
from logging.handlers import SMTPHandler
from openpyxl import load_workbook
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from dotenv import load_dotenv
load_dotenv()

# ---------------------------------------------------------------------------
# ✅ CONFIGURACÕES GLOBAIS  (segurança mantida conforme código original)
# ---------------------------------------------------------------------------
API_TOKEN = os.getenv("MILVUS_API_TOKEN")
SMTP_SERVER = os.getenv("SMTP_SERVER")
SMTP_PORT = int(os.getenv("SMTP_PORT"))
EMAIL_REMENTENTE = os.getenv("EMAIL_REMETENTE")
SENHA_REMENTENTE = os.getenv("SENHA_REMETENTE")
API_ENDPOINT = "https://apiintegracao.milvus.com.br/api/relatorio-atendimento/exporta"
DESTINATARIOS = [e.strip() for e in os.getenv("DESTINATARIOS_EMAIL").split(',')]
ASSUNTO_PADRAO = "Relatório de Atendimento Milvus"

EMAIL_DESTINATARIO_LOG = [e.strip() for e in os.getenv("EMAIL_DESTINATARIO_LOG").split(',')]
ASSUNTO_LOG_PADRAO = "Log de Execução - Script Relatório Milvus"
LOG_FILE_NAME = "relatorio_milvus.log" # Nome do arquivo de log
PYWHATKIT_DB_FILE = "PyWhatKit_DB.txt" # Nome do arquivo de DB do PyWhatKit

COLUNAS_PADRAO_A_EXCLUIR: List[str] = [
    "Categoria primária",
    "Categoria secundária",
    "contato",
    "Data chegada",
    "Data saida",
    "Data de finalização",
    "Descrição",
    "Mesa de trabalho",
    "Motivo de pausa do ticket",
    "setor",
    "Status",
    "Tipo de ticket",
    "Atendimento",
    "Atendimento em horário comercial?",
    "Atendimento externo?",
]
NOME_COLUNA_TECNICO = "Técnico"
NOME_COLUNA_TEMPO_ATENDIMENTO = "Tempo total de atendimento"

TECNICOS_A_IGNORAR = [t.strip() for t in os.getenv("TECNICOS_A_IGNORAR_LIST").split(',')]

LIMITE_MINIMO_HORAS = "04:00"
WHATSAPP_TECNICOS: Dict[str, str] = json.loads(os.getenv("WHATSAPP_TECNICOS_JSON"))

MESES_PT = [
    "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho", "Julho",
    "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro",
]

BASE_PASTA_RELATORIOS = Path("D:/Relatorios")

# ---------------------------------------------------------------------------
# 🔧 LOGGING
# ---------------------------------------------------------------------------
log_formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
root_logger = logging.getLogger() # Obter o logger raiz

# Configurar o console handler
console_handler = logging.StreamHandler()
console_handler.setFormatter(log_formatter)
console_handler.setLevel(logging.INFO)
root_logger.addHandler(console_handler)

# Configurar o file handler
file_handler = logging.FileHandler(LOG_FILE_NAME, encoding="utf-8")
file_handler.setFormatter(log_formatter)
file_handler.setLevel(logging.INFO)
root_logger.addHandler(file_handler)

# NÃO adiciona o SMTPHandler aqui. Ele será usado para enviar o log completo no final.
root_logger.setLevel(logging.INFO)

# ---------------------------------------------------------------------------
# 🌐 HTTP Session com Retry (Resiliência)
# ---------------------------------------------------------------------------


def _build_retry(total: int = 3, backoff_factor: float = 0.5) -> Retry:
    return Retry(
        total=total,
        backoff_factor=backoff_factor,
        status_forcelist=[500, 502, 503, 504],
        allowed_methods=["GET", "POST"],
        raise_on_status=False,
    )


def get_session() -> requests.Session:
    session = requests.Session()
    adapter = HTTPAdapter(max_retries=_build_retry())
    session.mount("https://", adapter)
    session.mount("http://", adapter)
    return session


SESSION = get_session()

# ---------------------------------------------------------------------------
# ⏲️ Funções utilitárias de tempo
# ---------------------------------------------------------------------------


def hhmm_to_total_minutes(hhmm_str: str) -> int:
    try:
        t = datetime.strptime(hhmm_str.strip(), "%H:%M")
        return t.hour * 60 + t.minute
    except Exception as exc:
        logging.error("Formato de tempo inválido '%s' (%s)", hhmm_str, exc)
        raise


def total_minutes_to_hhmm(total_minutes: int) -> str:
    hours, minutes = divmod(int(total_minutes), 60)
    return f"{hours:02d}:{minutes:02d}"

# ---------------------------------------------------------------------------
# 🌐 Conectividade
# ---------------------------------------------------------------------------


def verificar_conexao() -> bool:
    try:
        socket.create_connection(("8.8.8.8", 53), timeout=3)
        return True
    except OSError as exc:
        logging.error("❌ Erro de conexão de rede: %s", exc)
        return False

# ---------------------------------------------------------------------------
# 📡 API Milvus
# ---------------------------------------------------------------------------


def solicitar_dados_api(data_inicial: str, data_final: str) -> str:
    if not verificar_conexao():
        raise ConnectionError("Sem conexão à internet.")

    headers = {"Authorization": API_TOKEN, "Content-Type": "application/json"}
    body = {
        "filtro_body": {
            "data_inicial": data_inicial,
            "data_final": data_final,
            "tipo_arquivo": "csv",
        }
    }

    logging.info("📡 Solicitando dados da API (%s)…", data_inicial)
    start = time.perf_counter()
    resp = SESSION.post(API_ENDPOINT, headers=headers, json=body, timeout=30)
    duration = time.perf_counter() - start

    if resp.ok:
        logging.info("✅ Dados recebidos (%.1fs).", duration)
        return resp.text
    else:
        logging.error("❌ API HTTP %s – %s", resp.status_code, resp.text[:200])
        resp.raise_for_status()

# ---------------------------------------------------------------------------
# 🧹 Processamento CSV
# ---------------------------------------------------------------------------


def processar_csv(csv_content: str) -> pd.DataFrame:
    logging.info("🧠 Processando conteúdo CSV…")
    df = pd.read_csv(io.StringIO(csv_content), sep=";", encoding="utf-8", on_bad_lines="skip")

    cols_to_drop = [c for c in COLUNAS_PADRAO_A_EXCLUIR if c in df.columns]
    if cols_to_drop:
        df = df.drop(columns=cols_to_drop)
        # logging.info("Colunas removidas: %s", ", ".join(cols_to_drop)) # Removida a mensagem de log
    
    missing_cols = [c for c in (NOME_COLUNA_TECNICO, NOME_COLUNA_TEMPO_ATENDIMENTO) if c not in df.columns]
    if missing_cols:
        raise ValueError(f"Coluna(s) obrigatória(s) ausente(s): {missing_cols}")

    df = df.sort_values(by=NOME_COLUNA_TECNICO).reset_index(drop=True)
    return df

# ---------------------------------------------------------------------------
# 🆕  LINHA EM BRANCO + TOTAL POR TÉCNICO
# ---------------------------------------------------------------------------


def inserir_totais_por_tecnico(df: pd.DataFrame) -> pd.DataFrame:
    """
    Retorna um novo DataFrame no qual, para cada técnico, foi adicionada
    uma linha em branco logo após suas entradas contendo o total de horas
    desse técnico na coluna 'Tempo total de atendimento'.

    - A coluna 'Técnico' fica vazia na linha total.
    - Demais colunas também ficam vazias, exceto o total de horas.
    """
    blocos: List[pd.DataFrame] = []

    for tecnico, grupo in df.groupby(NOME_COLUNA_TECNICO, sort=False):
        blocos.append(grupo)

        # Soma as horas do grupo (HH:MM ➜ minutos ➜ HH:MM)
        total_min = (
            grupo[NOME_COLUNA_TEMPO_ATENDIMENTO]
            .str.split(":")
            .apply(lambda x: int(x[0]) * 60 + int(x[1]))
            .sum()
        )
        total_hhmm = total_minutes_to_hhmm(total_min)

        linha_total = {col: "" for col in df.columns}
        linha_total[NOME_COLUNA_TEMPO_ATENDIMENTO] = total_hhmm
        blocos.append(pd.DataFrame([linha_total]))

    return pd.concat(blocos, ignore_index=True)

# ---------------------------------------------------------------------------
# 🏁 Resumo por técnico
# ---------------------------------------------------------------------------


def calcular_soma_por_tecnico(df: pd.DataFrame) -> pd.DataFrame:
    df_work = df[~df[NOME_COLUNA_TECNICO].isin(TECNICOS_A_IGNORAR)].copy()
    td = pd.to_timedelta(df_work[NOME_COLUNA_TEMPO_ATENDIMENTO] + ":00")
    df_work["total_min"] = td.dt.total_seconds() // 60

    resumo = (
        df_work.groupby(NOME_COLUNA_TECNICO, sort=False)["total_min"].sum().reset_index()
    )
    resumo["Total Horas"] = resumo["total_min"].apply(total_minutes_to_hhmm)
    return resumo[[NOME_COLUNA_TECNICO, "Total Horas"]]

# ---------------------------------------------------------------------------
# ✉️  Geração do corpo de e-mail (texto + HTML)
# ---------------------------------------------------------------------------

def gerar_corpo_email(resumo: pd.DataFrame, data_str: str) -> tuple[str, str]:
    """Retorna (texto_plano, html)"""
    data_fmt = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")

    # ---------- Texto Plano ----------
    linhas_txt = "\n".join(f"{row[NOME_COLUNA_TECNICO]}: {row['Total Horas']}"
                           for _, row in resumo.iterrows())
    texto = (
        f"Prezados(as),\n\nConforme rotina diária, segue abaixo o informativo com o total de horas de atendimento registradas por cada colaborador no dia {data_fmt}:\n\n"
        f"{linhas_txt}\n\n Atenciosamente,\nEquipe JDB Tecnologia"
    )

    # ---------- HTML ----------
    linhas_html = "".join(
        f"<tr><td style='padding:4px 8px;border:1px solid #ccc'>{row[NOME_COLUNA_TECNICO]}</td>"
        f"<td style='padding:4px 8px;border:1px solid #ccc;text-align:center'>{row['Total Horas']}</td></tr>"
        for _, row in resumo.iterrows()
    )

    html = f"""
    <html>
      <body style="font-family:Arial,Helvetica,sans-serif;font-size:14px">
        <p>Prezados(as),<br><br>
           Conforme rotina diária, segue abaixo o informativo com o total de horas de atendimento registradas por cada colaborador no dia <b>{data_fmt}</b>:</p>
        <table style="border-collapse:collapse">
          <thead>
            <tr>
              <th style="padding:6px 10px;border:1px solid #ccc;background:#f0f0f0;text-align:left">Colaborador</th>
              <th style="padding:6px 10px;border:1px solid #ccc;background:#f0f0f0">Horas</th>
            </tr>
          </thead>
          <tbody>
            {linhas_html}
          </tbody>
        </table>
        <p style="margin-top:16px">Qualquer dúvida, estamos à disposição.<br>
           Atenciosamente,<br>Equipe JDB Tecnologia</p>
      </body>
    </html>
    """
    return texto, html

def enviar_email_com_anexo(caminho: Path, corpo_email: tuple[str, str],
                           destinatarios: List[str], assunto: str) -> None:
    texto, html = corpo_email
    msg = EmailMessage()
    msg["Subject"] = assunto
    msg["From"] = EMAIL_REMENTENTE
    msg["To"] = ", ".join(destinatarios)
    msg.set_content(texto)                 # parte texto plano
    msg.add_alternative(html, subtype="html")  # parte HTML

    with caminho.open("rb") as f:
        msg.add_attachment(f.read(), maintype="application",
                           subtype="octet-stream", filename=caminho.name)

    with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.login(EMAIL_REMENTENTE, SENHA_REMENTENTE)
        smtp.send_message(msg)
    logging.info("✅ E-mail enviado.")


# ---------------------------------------------------------------------------
# 📲 WhatsApp
# ---------------------------------------------------------------------------


def enviar_alerta_whatsapp(tecnico: str, horas: str, data_str: str) -> None:
    data_fmt = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
    if tecnico in TECNICOS_A_IGNORAR:
        return
    numero = WHATSAPP_TECNICOS.get(tecnico)
    if not numero:
        logging.warning("Número de WhatsApp não cadastrado para %s", tecnico)
        return

    mensagem = (
        f"Olá {tecnico}, tudo bem?\n\nVerificamos que o apontamento de horas do dia {data_fmt} está abaixo do esperado, foram registradas {horas} de atendimento no último dia útil.\n"
        "Pedimos, por gentileza, que revise as horas registradas e, se houver alguma pendência, que seja ajustada o quanto antes.\n"
        "⚠️O registro diário das horas é fundamental para garantir a transparência dos atendimentos e o correto acompanhamento das atividades da equipe.⚠️\n\n"
        "Agradecemos pela atenção e colaboração!\n"
        "*Equipe JDB Tecnologia*"
    )

    try:
        kit.sendwhatmsg_instantly(numero, mensagem, wait_time=30, tab_close=True)
        logging.info("WhatsApp enviado para %s", tecnico)
    except Exception as exc:
        logging.error("Erro ao enviar WhatsApp para %s (%s)", tecnico, exc)

# ---------------------------------------------------------------------------
# 📊 Excel mensal
# ---------------------------------------------------------------------------


def _mapear_planilha(ws) -> Tuple[Dict[str, int], Dict[int, int]]:
    tecnicos_col = {
        (ws.cell(row=2, column=col).value or "").split(" ")[0].lower(): col
        for col in range(2, ws.max_column + 1)
        if ws.cell(row=2, column=col).value is not None
    }
    dias_linha = {
        int(ws.cell(row=row, column=1).value): row
        for row in range(3, ws.max_row + 1)
        if str(ws.cell(row=row, column=1).value).isdigit()
    }
    return tecnicos_col, dias_linha


def atualizar_planilha_mensal(resumo: pd.DataFrame, data_ref: datetime) -> None:
    ano = data_ref.strftime("%Y")
    mes_num = data_ref.strftime("%m")
    mes_nome = MESES_PT[int(mes_num) - 1]
    pasta_mes = BASE_PASTA_RELATORIOS / ano / f"{mes_num}-{mes_nome}"
    planilha_path = pasta_mes / f"{mes_nome}.xlsx"

    logging.info("📓 Atualizando planilha %s…", planilha_path)

    if not planilha_path.exists():
        logging.error("Planilha %s não encontrada.", planilha_path)
        return

    wb = load_workbook(planilha_path)
    ws = wb.active

    map_tecnico, map_dia = _mapear_planilha(ws)
    dia = data_ref.day
    linha_dia = map_dia.get(dia)
    if not linha_dia:
        logging.error("Dia %s não encontrado na coluna A.", dia)
        return

    for _, row in resumo.iterrows():
        tecnico = row[NOME_COLUNA_TECNICO]
        if tecnico in TECNICOS_A_IGNORAR:
            continue
        col = map_tecnico.get(tecnico.split(" ")[0].lower())
        if col:
            ws.cell(row=linha_dia, column=col, value=row["Total Horas"])
        else:
            logging.warning("Técnico %s não encontrado no cabeçalho.", tecnico)

    wb.save(planilha_path)
    logging.info("Planilha mensal atualizada.")

# ---------------------------------------------------------------------------
# 💾 Persistência CSV
# ---------------------------------------------------------------------------


def salvar_csv(df: pd.DataFrame, data_ref: datetime) -> Path:
    ano = data_ref.strftime("%Y")
    mes_num = data_ref.strftime("%m")
    mes_nome = MESES_PT[int(mes_num) - 1]
    pasta = BASE_PASTA_RELATORIOS / ano / f"{mes_num}-{mes_nome}"
    pasta.mkdir(parents=True, exist_ok=True)
    caminho = pasta / f"{data_ref:%d}.csv"
    df.to_csv(caminho, sep=";", index=False, encoding="utf-8")
    logging.info("💾 CSV salvo em %s", caminho)
    return caminho

# ---------------------------------------------------------------------------
# 🚀 Função principal
# ---------------------------------------------------------------------------


def main() -> None:  # noqa: C901
    logging.info("Iniciando script Milvus…")
    start_run = time.perf_counter()

    # Define o último dia útil
    data_ref = datetime.now() - timedelta(days=1)
    while data_ref.weekday() >= 5: # 5 = Sábado, 6 = Domingo
        data_ref -= timedelta(days=1)

    data_str = data_ref.strftime("%Y-%m-%d")

    try:
        csv_raw = solicitar_dados_api(data_str, data_str)
        df_original = processar_csv(csv_raw)            # DataFrame “cru”
        df_para_csv = inserir_totais_por_tecnico(df_original)  # 👈 novo formato visível
        resumo = calcular_soma_por_tecnico(df_original)

        corpo_email = gerar_corpo_email(resumo, data_str) # Chamada para a função que foi adicionada
        
        # Alteração do assunto do e-mail para o formato desejado
        data_assunto = datetime.strptime(data_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        assunto_final = f"Resumo de Horas Trabalhadas – {data_assunto}"

        csv_path = salvar_csv(df_para_csv, data_ref)            # salva com linhas totais
        
        # Envia o e-mail com o relatório CSV anexado
        enviar_email_com_anexo(csv_path, corpo_email, DESTINATARIOS, assunto_final) # Usa assunto_final

        # Alertas WhatsApp
        limite_min = hhmm_to_total_minutes(LIMITE_MINIMO_HORAS)
        for _, row in resumo.iterrows():
            if hhmm_to_total_minutes(row["Total Horas"]) <= limite_min:
                enviar_alerta_whatsapp(row[NOME_COLUNA_TECNICO], row["Total Horas"], data_str)

        atualizar_planilha_mensal(resumo, data_ref)

        logging.info("✅ Script concluído em %.1fs.", time.perf_counter() - start_run)

    except Exception as exc:
        logging.critical("Erro fatal: %s", exc, exc_info=True)
    finally:
        # Fechar o file_handler para liberar o arquivo antes de tentar deletá-lo
        # Isso é crucial para resolver o erro WinError 32
        if file_handler in root_logger.handlers:
            file_handler.close()
            root_logger.removeHandler(file_handler)

        # Envia o arquivo de log completo por e-mail no final
        log_file_path = Path(LOG_FILE_NAME)
        if log_file_path.exists():
            try:
                # Cria uma nova mensagem de e-mail para o log
                msg_log = EmailMessage()
                msg_log["Subject"] = ASSUNTO_LOG_PADRAO
                msg_log["From"] = EMAIL_REMENTENTE
                msg_log["To"] = ", ".join(EMAIL_DESTINATARIO_LOG)
                msg_log.set_content("Log de execução do script 'relatorio_milvus'.") # Corpo simples

                with log_file_path.open("rb") as f:
                    msg_log.add_attachment(
                        f.read(),
                        maintype="text", # Tipo de conteúdo para um arquivo de texto
                        subtype="plain",
                        filename=log_file_path.name,
                    )
                
                with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as smtp:
                    smtp.starttls()
                    smtp.login(EMAIL_REMENTENTE, SENHA_REMENTENTE)
                    smtp.send_message(msg_log)
                # Note: logging.info aqui pode não aparecer no log_file_path se o handler foi fechado
                # Use print para garantir visibilidade no console
                print(f"✅ Log de execução enviado por e-mail para {', '.join(EMAIL_DESTINATARIO_LOG)}")

            except Exception as e:
                print(f"❌ Erro ao enviar arquivo de log por e-mail: {e}")
            finally:
                # Deleta o arquivo de log após tentar enviar
                try:
                    os.remove(log_file_path)
                    print(f"Arquivo de log '{LOG_FILE_NAME}' deletado com sucesso.")
                except OSError as e:
                    print(f"Erro ao deletar arquivo de log '{LOG_FILE_NAME}': [WinError 32] O arquivo já está sendo usado por outro processo: '{LOG_FILE_NAME}' - {e}")
        else:
            print(f"Arquivo de log '{LOG_FILE_NAME}' não encontrado para envio.")
        
        # --- NOVO: Deleta o arquivo PyWhatKit_DB.txt ---
        pywhatkit_db_path = Path(PYWHATKIT_DB_FILE)
        if pywhatkit_db_path.exists():
            try:
                os.remove(pywhatkit_db_path)
                print(f"Arquivo '{PYWHATKIT_DB_FILE}' deletado com sucesso.")
            except OSError as e:
                print(f"Erro ao deletar arquivo '{PYWHATKIT_DB_FILE}': {e}")


if __name__ == "__main__":
    main()